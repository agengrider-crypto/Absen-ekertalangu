from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import logging
import base64
import io
import re
import secrets
import hashlib
import hmac
import uuid
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Annotated

import bcrypt
import jwt
import qrcode
from bson import ObjectId
from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File
from starlette.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, BeforeValidator, ConfigDict

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000").rstrip("/")
# Session lifetime (masa percobaan): sesi bertahan 365 hari agar tetap login saat refresh web
SESSION_DAYS = 365
SESSION_MAX_AGE = SESSION_DAYS * 24 * 60 * 60  # detik
VALID_ROLES = ["admin", "pengurus", "peserta"]
EDUCATION_OPTIONS = ["TK", "SD", "SMP", "SMA", "D1", "D2", "D3", "D4", "S1", "S2", "S3"]
MUBALIGH_OPTIONS = ["belum", "sudah"]
GENDER_OPTIONS = ["L", "P"]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ---------------------------------------------------------------------------
# Helpers: password + jwt
# ---------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def create_access_token(user_id: str, ver: int = 0) -> str:
    payload = {"sub": user_id, "ver": ver, "type": "access",
               "exp": datetime.now(timezone.utc) + timedelta(days=SESSION_DAYS)}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str, ver: int = 0) -> str:
    payload = {"sub": user_id, "ver": ver, "type": "refresh",
               "exp": datetime.now(timezone.utc) + timedelta(days=SESSION_DAYS)}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, user_id: str, ver: int):
    response.set_cookie("access_token", create_access_token(user_id, ver), httponly=True,
                        secure=True, samesite="none", max_age=SESSION_MAX_AGE, path="/")
    response.set_cookie("refresh_token", create_refresh_token(user_id, ver), httponly=True,
                        secure=True, samesite="none", max_age=SESSION_MAX_AGE, path="/")

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
PyObjectId = Annotated[str, BeforeValidator(str)]

def _derive_gender(user: dict) -> Optional[str]:
    g = user.get("gender")
    if g in GENDER_OPTIONS:
        return g
    av = user.get("avatar_gender")
    if av == "female":
        return "P"
    if av == "male":
        return "L"
    return None

def public_user(user: dict, include_photo: bool = False) -> dict:
    data = {
        "id": str(user["_id"]),
        "name": user.get("name"),
        "email": user.get("email"),
        "username": user.get("username"),
        "phone": user.get("phone"),
        "whatsapp": user.get("whatsapp"),
        "dob": user.get("dob"),
        "birthplace": user.get("birthplace"),
        "address": user.get("address"),
        "gender": _derive_gender(user),
        "education": user.get("education"),
        "mubaligh": user.get("mubaligh"),
        "kelompok_id": user.get("kelompok_id"),
        "roles": user.get("roles", []),
        "status": user.get("status", "active"),
        "source": user.get("source", "admin"),
        "avatar_gender": user.get("avatar_gender", "male"),
        "needs_completion": bool(user.get("needs_completion", False)),
        "has_photo": bool(user.get("photo")),
        "created_at": user.get("created_at"),
    }
    if include_photo:
        data["photo"] = user.get("photo")
    return data

class LoginInput(BaseModel):
    identifier: str
    password: str

class RegisterInput(BaseModel):
    token: str
    name: str
    phone: str
    email: str
    dob: str
    address: str
    password: str
    avatar_gender: Optional[str] = "male"

class ActivationComplete(BaseModel):
    user_id: str
    phone: str
    email: Optional[str] = None
    dob: str
    address: Optional[str] = None
    password: str
    gender: Optional[str] = None
    avatar_gender: Optional[str] = "male"

class AdminCreateUser(BaseModel):
    name: str
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    dob: Optional[str] = None
    birthplace: Optional[str] = None
    address: Optional[str] = None
    gender: Optional[str] = None
    education: Optional[str] = None
    mubaligh: Optional[str] = None
    kelompok_id: Optional[str] = None
    roles: List[str] = ["peserta"]
    password: Optional[str] = None

class PendingEntry(BaseModel):
    name: str
    dob: Optional[str] = None

class AdminPendingNames(BaseModel):
    entries: List[PendingEntry]

class SelfResetInput(BaseModel):
    phone: str
    dob: str
    new_password: str

# --- Fase 2: Kelompok, Peserta management ---
class KelompokInput(BaseModel):
    name: str
    description: Optional[str] = None

class BulkEntry(BaseModel):
    name: str
    gender: Optional[str] = None
    birthplace: Optional[str] = None
    dob: Optional[str] = None
    phone: Optional[str] = None

class BulkCreateInput(BaseModel):
    entries: List[BulkEntry]
    kelompok_id: Optional[str] = None

class BulkDeleteInput(BaseModel):
    ids: List[str]

class PesertaUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    dob: Optional[str] = None
    birthplace: Optional[str] = None
    address: Optional[str] = None
    gender: Optional[str] = None
    education: Optional[str] = None
    mubaligh: Optional[str] = None
    photo: Optional[str] = None
    roles: Optional[List[str]] = None
    status: Optional[str] = None
    kelompok_id: Optional[str] = None
    needs_completion: Optional[bool] = None

class MoveInput(BaseModel):
    kelompok_id: Optional[str] = None
    keterangan: Optional[str] = None


class PhotoInput(BaseModel):
    photo: Optional[str] = None


class FeedbackInput(BaseModel):
    name: Optional[str] = None
    message: str


class SelfAbsenInput(BaseModel):
    user_id: str

# ---------------------------------------------------------------------------
# Auth dependency
# ---------------------------------------------------------------------------
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Tipe token tidak valid")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Pengguna tidak ditemukan")
        if payload.get("ver", 0) != user.get("token_version", 0):
            raise HTTPException(status_code=401, detail="Sesi telah berakhir")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token kadaluarsa")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if "admin" not in user.get("roles", []):
        raise HTTPException(status_code=403, detail="Akses khusus admin")
    return user

async def require_staff(user: dict = Depends(get_current_user)) -> dict:
    roles = user.get("roles", [])
    if "admin" not in roles and "pengurus" not in roles:
        raise HTTPException(status_code=403, detail="Akses khusus admin/pengurus")
    return user

def is_admin(user: dict) -> bool:
    return "admin" in (user.get("roles", []) or [])

def normalize_dob(value) -> Optional[str]:
    """Return YYYY-MM-DD or None. Accepts date/datetime and common ID string formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):  # date object from openpyxl
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def normalize_gender(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("l", "laki-laki", "laki", "pria", "male", "m", "lk"):
        return "L"
    if s in ("p", "perempuan", "wanita", "female", "f", "pr"):
        return "P"
    return None

def build_pending_doc(name: str, dob: Optional[str]) -> dict:
    return {
        "name": name, "email": None, "username": None, "phone": None,
        "dob": normalize_dob(dob), "address": None, "roles": ["peserta"],
        "source": "admin_import", "avatar_gender": "male", "password_hash": None,
        "status": "pending", "token_version": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

async def log_activity(actor: Optional[dict], action: str, detail: str = "", target: str = ""):
    """Persist an audit log entry. Never raises."""
    try:
        await db.activity_logs.insert_one({
            "_id": str(uuid.uuid4()),
            "actor_id": str(actor["_id"]) if actor else None,
            "actor_name": actor.get("name") if actor else "Sistem",
            "action": action,
            "detail": detail,
            "target": target,
            "at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as exc:  # pragma: no cover
        logger.warning("Gagal mencatat log: %s", exc)

# ---------------------------------------------------------------------------
# QR helpers (server-side generated & cached)
# ---------------------------------------------------------------------------
def make_qr_data_url(data: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=10, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#0D5C3A", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()

async def get_or_create_public_qr() -> dict:
    doc = await db.app_settings.find_one({"_id": "public_qr"})
    if not doc:
        token = secrets.token_urlsafe(12)
        link = f"{FRONTEND_URL}/register?token={token}"
        doc = {"_id": "public_qr", "token": token, "link": link,
               "image": make_qr_data_url(link), "created_at": datetime.now(timezone.utc).isoformat()}
        await db.app_settings.insert_one(doc)
    return doc

# ---------------------------------------------------------------------------
# Brute force helpers
# ---------------------------------------------------------------------------
async def is_locked(identifier: str) -> bool:
    since = datetime.now(timezone.utc) - timedelta(minutes=15)
    count = await db.login_attempts.count_documents(
        {"identifier": identifier, "at": {"$gte": since.isoformat()}})
    return count >= 5

# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@api_router.get("/")
async def root():
    return {"message": "E-KERTALANGU API"}

@api_router.post("/auth/login")
async def login(body: LoginInput, request: Request, response: Response):
    ident = body.identifier.strip().lower()
    user = await db.users.find_one({"$or": [
        {"email": ident}, {"username": ident}, {"phone": body.identifier.strip()}]})
    lock_key = str(user["_id"]) if user else ident
    if await is_locked(lock_key):
        raise HTTPException(status_code=429, detail="Terlalu banyak percobaan. Coba lagi dalam 15 menit.")

    ok = user and user.get("password_hash") and verify_password(body.password, user["password_hash"])
    if not ok:
        await db.login_attempts.insert_one({
            "identifier": lock_key, "email": user["email"] if user else ident,
            "at": datetime.now(timezone.utc).isoformat()})
        raise HTTPException(status_code=401, detail="Akun atau kata sandi salah")
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Akun belum aktif. Silakan aktivasi terlebih dahulu.")

    await db.login_attempts.delete_many({"identifier": lock_key})
    set_auth_cookies(response, str(user["_id"]), user.get("token_version", 0))
    await log_activity(user, "login", f"Login berhasil sebagai {', '.join(user.get('roles', []))}")
    return public_user(user)

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Keluar berhasil"}


def parse_object_id(value: str) -> ObjectId:
    try:
        return ObjectId(value)
    except Exception:
        raise HTTPException(status_code=400, detail="ID pengguna tidak valid")

@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return public_user(user)


@api_router.get("/me/photo")
async def get_my_photo(user: dict = Depends(get_current_user)):
    return {"photo": user.get("photo")}


@api_router.post("/me/photo")
async def set_my_photo(body: PhotoInput, user: dict = Depends(get_current_user)):
    photo = (body.photo or "").strip() or None
    if photo and not photo.startswith("data:image/"):
        raise HTTPException(status_code=400, detail="Format foto tidak valid (harus JPG/PNG).")
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"photo": photo}})
    return {"has_photo": bool(photo)}

@api_router.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Tidak ada sesi")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Token tidak valid")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user or payload.get("ver", 0) != user.get("token_version", 0):
            raise HTTPException(status_code=401, detail="Sesi telah berakhir")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")
    response.set_cookie("access_token", create_access_token(str(user["_id"]), user.get("token_version", 0)),
                        httponly=True, secure=True, samesite="none", max_age=SESSION_MAX_AGE, path="/")
    return public_user(user)

@api_router.post("/auth/register")
async def register(body: RegisterInput, response: Response):
    qr = await db.app_settings.find_one({"_id": "public_qr"})
    if not qr or body.token != qr["token"]:
        raise HTTPException(status_code=400, detail="Kode QR pendaftaran tidak valid atau kadaluarsa")

    email = body.email.strip().lower()
    phone = body.phone.strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email sudah terdaftar")
    if await db.users.find_one({"phone": phone}):
        raise HTTPException(status_code=409, detail="Nomor HP sudah terdaftar")

    doc = {
        "name": body.name.strip(),
        "email": email,
        "username": None,
        "phone": phone,
        "dob": normalize_dob(body.dob) or body.dob.strip(),
        "address": body.address.strip(),
        "password_hash": hash_password(body.password),
        "roles": ["peserta"],
        "status": "active",
        "source": "qr_public",
        "avatar_gender": body.avatar_gender or "male",
        "token_version": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    set_auth_cookies(response, str(res.inserted_id), 0)
    return public_user(doc)

@api_router.get("/activation/search")
async def activation_search(q: str = ""):
    q = q.strip()
    if len(q) < 2:
        return []
    cur = db.users.find(
        {"status": "pending", "name": {"$regex": re.escape(q), "$options": "i"}}
    ).sort("name", 1).limit(20)
    results = await cur.to_list(20)
    return [{"id": str(u["_id"]), "name": u["name"], "requires_dob": bool(u.get("dob"))} for u in results]

@api_router.post("/activation/complete")
async def activation_complete(body: ActivationComplete, response: Response):
    user = await db.users.find_one({"_id": parse_object_id(body.user_id)})
    if not user or user.get("status") != "pending":
        raise HTTPException(status_code=404, detail="Data peserta tidak ditemukan atau sudah aktif")

    submitted_dob = normalize_dob(body.dob)
    if user.get("dob") and submitted_dob != user["dob"]:
        raise HTTPException(status_code=403, detail="Tanggal lahir tidak sesuai data yang didaftarkan pengurus.")

    gender = normalize_gender(body.gender)
    if not gender:
        raise HTTPException(status_code=400, detail="Jenis kelamin wajib diisi (Laki-laki / Perempuan).")

    phone = body.phone.strip()
    if not phone:
        raise HTTPException(status_code=400, detail="Nomor HP wajib diisi.")
    email = (body.email or "").strip().lower() or None
    address = (body.address or "").strip() or None

    if email:
        dup_email = await db.users.find_one({"email": email, "_id": {"$ne": user["_id"]}})
        if dup_email:
            raise HTTPException(status_code=409, detail="Email sudah terdaftar")
    dup_phone = await db.users.find_one({"phone": phone, "_id": {"$ne": user["_id"]}})
    if dup_phone:
        raise HTTPException(status_code=409, detail="Nomor HP sudah terdaftar")

    await db.users.update_one({"_id": user["_id"]}, {"$set": {
        "phone": phone, "email": email, "dob": submitted_dob or body.dob.strip(),
        "address": address, "password_hash": hash_password(body.password),
        "status": "active", "gender": gender,
        "avatar_gender": "female" if gender == "P" else "male"}})
    user = await db.users.find_one({"_id": user["_id"]})
    set_auth_cookies(response, str(user["_id"]), user.get("token_version", 0))
    await log_activity(user, "aktivasi_akun", "Akun diaktivasi & data dilengkapi")
    return public_user(user)

@api_router.post("/auth/self-reset")
async def self_reset(body: SelfResetInput):
    phone = body.phone.strip()
    dob = normalize_dob(body.dob) or body.dob.strip()
    user = await db.users.find_one({"phone": phone, "dob": dob})
    if not user:
        raise HTTPException(status_code=404, detail="Data tidak cocok. Periksa Nomor HP & tanggal lahir Anda.")
    if set(user.get("roles", [])) != {"peserta"}:
        raise HTTPException(status_code=403, detail="Reset mandiri hanya untuk akun Peserta. Akun dengan peran lain hubungi administrator.")
    await db.users.update_one({"_id": user["_id"]}, {
        "$set": {"password_hash": hash_password(body.new_password), "status": "active"},
        "$inc": {"token_version": 1}})
    await db.login_attempts.delete_many({"email": user["email"]})
    return {"message": "Kata sandi berhasil diperbarui. Silakan login."}

# ---- QR ----
@api_router.get("/qr/public")
async def public_qr():
    doc = await get_or_create_public_qr()
    return {"link": doc["link"], "image": doc["image"], "token": doc["token"]}

# ---- Admin ----
@api_router.get("/admin/users")
async def admin_users(admin: dict = Depends(require_staff)):
    users = await db.users.find().sort("created_at", -1).to_list(1000)
    return [public_user(u) for u in users]


@api_router.get("/admin/users/{user_id}/photo")
async def admin_user_photo(user_id: str, admin: dict = Depends(require_staff)):
    user = await db.users.find_one({"_id": parse_object_id(user_id)})
    photo = (user or {}).get("photo")
    if not user or not photo or not isinstance(photo, str) or "," not in photo:
        raise HTTPException(status_code=404, detail="Foto tidak ditemukan")
    try:
        header, b64 = photo.split(",", 1)
        media = "image/jpeg"
        if header.startswith("data:") and ";" in header:
            media = header[5:].split(";", 1)[0] or media
        raw = base64.b64decode(b64)
    except Exception:
        raise HTTPException(status_code=404, detail="Foto tidak valid")
    return Response(content=raw, media_type=media,
                    headers={"Cache-Control": "private, max-age=60"})

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if user_id == str(admin["_id"]):
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri")
    target = await db.users.find_one({"_id": parse_object_id(user_id)})
    res = await db.users.delete_one({"_id": parse_object_id(user_id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    await log_activity(admin, "hapus_akun", f"Menghapus akun '{target.get('name') if target else user_id}'", user_id)
    return {"message": "Pengguna dihapus"}

class RoleUpdate(BaseModel):
    roles: List[str]

@api_router.patch("/admin/users/{user_id}/roles")
async def admin_update_roles(user_id: str, body: RoleUpdate, admin: dict = Depends(require_admin)):
    roles = [r for r in body.roles if r in VALID_ROLES]
    if not roles:
        raise HTTPException(status_code=400, detail="Minimal satu peran valid")
    oid = parse_object_id(user_id)
    await db.users.update_one({"_id": oid}, {"$set": {"roles": roles}})
    user = await db.users.find_one({"_id": oid})
    return public_user(user)

@api_router.post("/admin/users")
async def admin_create_user(body: AdminCreateUser, admin: dict = Depends(require_staff)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    roles = [r for r in body.roles if r in VALID_ROLES] or ["peserta"]
    email = body.email.strip().lower() if body.email else None
    phone = body.phone.strip() if body.phone else None
    if email and await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email sudah terdaftar")
    if phone and await db.users.find_one({"phone": phone}):
        raise HTTPException(status_code=409, detail="Nomor HP sudah terdaftar")

    doc = {
        "name": name, "email": email, "username": None, "phone": phone,
        "whatsapp": body.whatsapp.strip() if body.whatsapp else None,
        "dob": normalize_dob(body.dob),
        "birthplace": body.birthplace.strip() if body.birthplace else None,
        "address": body.address.strip() if body.address else None,
        "gender": normalize_gender(body.gender),
        "education": body.education if body.education in EDUCATION_OPTIONS else None,
        "mubaligh": body.mubaligh if body.mubaligh in MUBALIGH_OPTIONS else None,
        "kelompok_id": body.kelompok_id or None,
        "roles": roles, "source": "admin",
        "avatar_gender": "female" if normalize_gender(body.gender) == "P" else "male",
        "token_version": 0, "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if body.password:
        doc["password_hash"] = hash_password(body.password)
        doc["status"] = "active"
    else:
        doc["password_hash"] = None
        doc["status"] = "pending"
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    await log_activity(admin, "buat_akun", f"Membuat akun '{name}' ({', '.join(roles)})", str(res.inserted_id))
    return public_user(doc)

@api_router.post("/admin/users/pending")
async def admin_create_pending(body: AdminPendingNames, admin: dict = Depends(require_staff)):
    created, skipped, invalid_dates = [], [], []
    for entry in body.entries:
        name = entry.name.strip()
        if not name:
            continue
        existing = await db.users.find_one({"name": name, "status": "pending"})
        if existing:
            skipped.append(name)
            continue
        if entry.dob and str(entry.dob).strip() and normalize_dob(entry.dob) is None:
            invalid_dates.append({"name": name, "value": str(entry.dob).strip()})
        doc = build_pending_doc(name, entry.dob)
        res = await db.users.insert_one(doc)
        doc["_id"] = res.inserted_id
        created.append(public_user(doc))
    return {"created": created, "count": len(created), "skipped": skipped, "invalid_dates": invalid_dates}

@api_router.get("/admin/import-template")
async def import_template(admin: dict = Depends(require_staff)):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peserta"
    ws.append(["Nama", "Tanggal Lahir"])
    ws.append(["Budi Santoso", "17-08-1970"])
    ws.append(["Siti Aminah", "02-05-1965"])
    ws.append(["Ahmad Fauzi", ""])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_peserta_ekertalangu.xlsx"},
    )

@api_router.post("/admin/users/import")
async def admin_import_users(file: UploadFile = File(...), admin: dict = Depends(require_staff)):
    fname = (file.filename or "").lower()
    data = await file.read()
    rows = []  # list of (name, dob)
    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            import csv
            text = data.decode("utf-8-sig", errors="ignore")
            for cols in csv.reader(io.StringIO(text)):
                if cols:
                    rows.append((cols[0], cols[1] if len(cols) > 1 else None))
        elif fname.endswith(".xlsx") or fname.endswith(".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and r[0] is not None:
                    rows.append((r[0], r[1] if len(r) > 1 else None))
            wb.close()
        else:
            raise HTTPException(status_code=400, detail="Format tidak didukung. Gunakan .xlsx atau .csv")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Gagal membaca file. Pastikan format .xlsx atau .csv benar.")

    created, skipped, invalid_dates, flagged = [], [], [], []
    for name_val, dob_val in rows:
        name = str(name_val).strip()
        if not name or name.lower() in ("nama", "name", "nama lengkap", "nama peserta"):
            continue
        # Nama kembar: tetap dibuat, tandai perlu dilengkapi (jangan dilewati)
        duplicate = await db.users.find_one({"name": name})
        if dob_val is not None and str(dob_val).strip() and normalize_dob(dob_val) is None:
            invalid_dates.append({"name": name, "value": str(dob_val).strip()})
        doc = build_pending_doc(name, dob_val)
        if duplicate:
            doc["needs_completion"] = True
            flagged.append(name)
        res = await db.users.insert_one(doc)
        created.append(name)
    await log_activity(admin, "impor_peserta",
                       f"Impor {len(created)} peserta dari file ({len(flagged)} nama kembar ditandai)")
    return {"count": len(created), "skipped": skipped, "created_names": created,
            "invalid_dates": invalid_dates, "flagged": flagged}

# ---------------------------------------------------------------------------
# Fase 2: Kelompok / Majelis
# ---------------------------------------------------------------------------
@api_router.get("/admin/kelompok")
async def list_kelompok(admin: dict = Depends(require_staff)):
    items = await db.kelompoks.find().sort("name", 1).to_list(500)
    counts = await db.users.aggregate([
        {"$match": {"kelompok_id": {"$ne": None}}},
        {"$group": {"_id": "$kelompok_id", "n": {"$sum": 1}}},
    ]).to_list(1000)
    count_map = {c["_id"]: c["n"] for c in counts}
    return [{
        "id": k["_id"], "name": k.get("name"), "description": k.get("description"),
        "member_count": count_map.get(k["_id"], 0), "created_at": k.get("created_at"),
    } for k in items]

@api_router.post("/admin/kelompok")
async def create_kelompok(body: KelompokInput, admin: dict = Depends(require_admin)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama kelompok wajib diisi")
    if await db.kelompoks.find_one({"name": name}):
        raise HTTPException(status_code=409, detail="Nama kelompok sudah ada")
    doc = {"_id": str(uuid.uuid4()), "name": name,
           "description": body.description.strip() if body.description else None,
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.kelompoks.insert_one(doc)
    await log_activity(admin, "buat_kelompok", f"Membuat kelompok '{name}'", doc["_id"])
    return {"id": doc["_id"], "name": doc["name"], "description": doc["description"], "member_count": 0}

@api_router.patch("/admin/kelompok/{kelompok_id}")
async def update_kelompok(kelompok_id: str, body: KelompokInput, admin: dict = Depends(require_admin)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama kelompok wajib diisi")
    dup = await db.kelompoks.find_one({"name": name, "_id": {"$ne": kelompok_id}})
    if dup:
        raise HTTPException(status_code=409, detail="Nama kelompok sudah ada")
    res = await db.kelompoks.update_one({"_id": kelompok_id}, {"$set": {
        "name": name, "description": body.description.strip() if body.description else None}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kelompok tidak ditemukan")
    return {"id": kelompok_id, "name": name}

@api_router.delete("/admin/kelompok/{kelompok_id}")
async def delete_kelompok(kelompok_id: str, admin: dict = Depends(require_admin)):
    res = await db.kelompoks.delete_one({"_id": kelompok_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kelompok tidak ditemukan")
    await db.users.update_many({"kelompok_id": kelompok_id}, {"$set": {"kelompok_id": None}})
    await log_activity(admin, "hapus_kelompok", f"Menghapus kelompok {kelompok_id}", kelompok_id)
    return {"message": "Kelompok dihapus"}

# ---------------------------------------------------------------------------
# Fase 2: Peserta detail + update + bulk + reset + move
# ---------------------------------------------------------------------------
@api_router.get("/admin/users/{user_id}")
async def admin_user_detail(user_id: str, admin: dict = Depends(require_staff)):
    user = await db.users.find_one({"_id": parse_object_id(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    data = public_user(user, include_photo=True)
    if user.get("kelompok_id"):
        k = await db.kelompoks.find_one({"_id": user["kelompok_id"]})
        data["kelompok_name"] = k.get("name") if k else None
    return data

@api_router.patch("/admin/users/{user_id}")
async def admin_update_user(user_id: str, body: PesertaUpdate, admin: dict = Depends(require_staff)):
    oid = parse_object_id(user_id)
    user = await db.users.find_one({"_id": oid})
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    updates = {}
    if body.name is not None and body.name.strip():
        updates["name"] = body.name.strip()
    if body.email is not None:
        email = body.email.strip().lower() or None
        if email and await db.users.find_one({"email": email, "_id": {"$ne": oid}}):
            raise HTTPException(status_code=409, detail="Email sudah terdaftar")
        updates["email"] = email
    if body.phone is not None:
        phone = body.phone.strip() or None
        if phone and await db.users.find_one({"phone": phone, "_id": {"$ne": oid}}):
            raise HTTPException(status_code=409, detail="Nomor HP sudah terdaftar")
        updates["phone"] = phone
    if body.whatsapp is not None:
        updates["whatsapp"] = body.whatsapp.strip() or None
    if body.dob is not None:
        updates["dob"] = normalize_dob(body.dob)
    if body.birthplace is not None:
        updates["birthplace"] = body.birthplace.strip() or None
    if body.address is not None:
        updates["address"] = body.address.strip() or None
    if body.gender is not None:
        g = normalize_gender(body.gender)
        updates["gender"] = g
        updates["avatar_gender"] = "female" if g == "P" else "male"
    if body.education is not None:
        updates["education"] = body.education if body.education in EDUCATION_OPTIONS else None
    if body.mubaligh is not None:
        updates["mubaligh"] = body.mubaligh if body.mubaligh in MUBALIGH_OPTIONS else None
    if body.photo is not None:
        updates["photo"] = body.photo or None
    if body.kelompok_id is not None:
        updates["kelompok_id"] = body.kelompok_id or None
    if body.needs_completion is not None:
        updates["needs_completion"] = bool(body.needs_completion)
    if body.roles is not None:
        roles = [r for r in body.roles if r in VALID_ROLES]
        if not roles:
            raise HTTPException(status_code=400, detail="Minimal satu peran valid")
        updates["roles"] = roles
    if body.status is not None:
        if body.status not in ("active", "nonaktif", "pending"):
            raise HTTPException(status_code=400, detail="Status tidak valid")
        if body.status == "active" and not user.get("password_hash"):
            raise HTTPException(status_code=400, detail="Akun belum punya kata sandi. Tetapkan sandi dulu (reset).")
        updates["status"] = body.status
    if not updates:
        return public_user(user, include_photo=True)
    await db.users.update_one({"_id": oid}, {"$set": updates})
    user = await db.users.find_one({"_id": oid})
    await log_activity(admin, "ubah_peserta", f"Mengubah data '{user.get('name')}'", user_id)
    return public_user(user, include_photo=True)

@api_router.post("/admin/users/bulk")
async def admin_bulk_create(body: BulkCreateInput, admin: dict = Depends(require_staff)):
    created, invalid_dates, flagged = [], [], []
    for entry in body.entries:
        name = (entry.name or "").strip()
        if not name:
            continue
        duplicate = await db.users.find_one({"name": name})
        if entry.dob and str(entry.dob).strip() and normalize_dob(entry.dob) is None:
            invalid_dates.append({"name": name, "value": str(entry.dob).strip()})
        g = normalize_gender(entry.gender)
        doc = {
            "name": name, "email": None, "username": None,
            "phone": entry.phone.strip() if entry.phone else None, "whatsapp": None,
            "dob": normalize_dob(entry.dob),
            "birthplace": entry.birthplace.strip() if entry.birthplace else None,
            "address": None, "gender": g,
            "avatar_gender": "female" if g == "P" else "male",
            "education": None, "mubaligh": None,
            "kelompok_id": body.kelompok_id or None,
            "roles": ["peserta"], "source": "admin_bulk", "password_hash": None,
            "status": "pending", "token_version": 0,
            "needs_completion": bool(duplicate),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        if duplicate:
            flagged.append(name)
        await db.users.insert_one(doc)
        created.append(name)
    await log_activity(admin, "tambah_bulk", f"Menambah {len(created)} peserta (bulk)")
    return {"count": len(created), "flagged": flagged, "invalid_dates": invalid_dates}

@api_router.post("/admin/users/bulk-delete")
async def admin_bulk_delete(body: BulkDeleteInput, admin: dict = Depends(require_admin)):
    ids = [i for i in body.ids if i != str(admin["_id"])]
    oids = []
    for i in ids:
        try:
            oids.append(ObjectId(i))
        except Exception:
            continue
    if not oids:
        raise HTTPException(status_code=400, detail="Tidak ada ID valid untuk dihapus")
    res = await db.users.delete_many({"_id": {"$in": oids}})
    await log_activity(admin, "hapus_massal", f"Menghapus {res.deleted_count} akun sekaligus")
    return {"deleted": res.deleted_count}

@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, admin: dict = Depends(require_staff)):
    oid = parse_object_id(user_id)
    user = await db.users.find_one({"_id": oid})
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    dob = normalize_dob(user.get("dob"))
    if not dob:
        raise HTTPException(status_code=400, detail="Tanggal lahir belum diisi. Lengkapi dulu untuk membuat sandi ddmmyyyy.")
    y, m, d = dob.split("-")
    new_pw = f"{d}{m}{y}"  # ddmmyyyy
    set_fields = {"password_hash": hash_password(new_pw)}
    if user.get("status") == "pending":
        set_fields["status"] = "active"
    await db.users.update_one({"_id": oid}, {"$set": set_fields, "$inc": {"token_version": 1}})
    await log_activity(admin, "reset_sandi", f"Reset sandi '{user.get('name')}' (format ddmmyyyy)", user_id)
    return {"password": new_pw, "message": "Kata sandi direset ke format tanggal lahir (ddmmyyyy)."}

@api_router.post("/admin/users/{user_id}/move")
async def admin_move_kelompok(user_id: str, body: MoveInput, admin: dict = Depends(require_staff)):
    oid = parse_object_id(user_id)
    user = await db.users.find_one({"_id": oid})
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    kname = None
    if body.kelompok_id:
        k = await db.kelompoks.find_one({"_id": body.kelompok_id})
        if not k:
            raise HTTPException(status_code=404, detail="Kelompok tujuan tidak ditemukan")
        kname = k.get("name")
    await db.users.update_one({"_id": oid}, {"$set": {"kelompok_id": body.kelompok_id or None}})
    ket = (body.keterangan or "").strip()
    detail = f"Pindah '{user.get('name')}' ke kelompok {kname or '-'}"
    if ket:
        detail += f" — Ket: {ket}"
    await log_activity(admin, "pindah_sambung", detail, user_id)
    return {"message": "Pindah sambung berhasil", "kelompok_id": body.kelompok_id or None,
            "kelompok_name": kname, "keterangan": ket or None}

# ---------------------------------------------------------------------------
# Fase 2: Log Aktivitas
# ---------------------------------------------------------------------------
@api_router.get("/admin/logs")
async def admin_logs(admin: dict = Depends(require_staff), limit: int = 100, action: str = ""):
    query = {}
    if action.strip():
        query["action"] = action.strip()
    limit = max(1, min(limit, 500))
    logs = await db.activity_logs.find(query).sort("at", -1).to_list(limit)
    return [{
        "id": l["_id"], "actor_id": l.get("actor_id"), "actor_name": l.get("actor_name"),
        "action": l.get("action"), "detail": l.get("detail"), "target": l.get("target"),
        "at": l.get("at"),
    } for l in logs]

# ===========================================================================
# FASE 2 — KEGIATAN + ABSENSI + DASHBOARD + LAPORAN (Admin)
# ===========================================================================
import asyncio

WITA = timezone(timedelta(hours=8))
KEGIATAN_TYPES = ["rutin", "khusus", "asad"]
ABSEN_STATUS = ["hadir", "izin", "alpha"]
PESERTA_QUERY = {"roles": "peserta", "status": "active"}
SHARE_EXPIRE_DAYS = 7


def now_wita() -> datetime:
    return datetime.now(WITA)


def kegiatan_end_dt(date_str: str, end_time: str):
    try:
        return datetime.strptime(f"{date_str} {end_time}", "%Y-%m-%d %H:%M").replace(tzinfo=WITA)
    except Exception:
        return None


class KegiatanInput(BaseModel):
    name: str
    type: str = "rutin"
    date: str  # YYYY-MM-DD (WITA)
    start_time: str  # HH:MM (WITA, 24h)
    end_time: str
    teacher: Optional[str] = None
    material: Optional[str] = None
    location: Optional[str] = None
    recurring: bool = False


class KegiatanUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    teacher: Optional[str] = None
    material: Optional[str] = None
    location: Optional[str] = None


class AbsenInput(BaseModel):
    user_id: str
    status: str  # hadir | izin | alpha


def serialize_kegiatan(k: dict, counts: Optional[dict] = None) -> dict:
    d = {
        "id": k["_id"],
        "name": k.get("name"),
        "type": k.get("type", "rutin"),
        "date": k.get("date"),
        "start_time": k.get("start_time"),
        "end_time": k.get("end_time"),
        "teacher": k.get("teacher"),
        "material": k.get("material"),
        "location": k.get("location"),
        "recurring": bool(k.get("recurring")),
        "status": k.get("status", "open"),
        "closed_at": k.get("closed_at"),
        "auto_closed": bool(k.get("auto_closed", False)),
        "share_token": k.get("share_token"),
        "share_expires_at": k.get("share_expires_at"),
        "created_at": k.get("created_at"),
    }
    if counts is not None:
        d["counts"] = counts
    return d


async def count_peserta_total() -> int:
    return await db.users.count_documents(PESERTA_QUERY)


async def kegiatan_counts(kegiatan_id: str, total: Optional[int] = None) -> dict:
    if total is None:
        total = await count_peserta_total()
    hadir = await db.absensis.count_documents({"kegiatan_id": kegiatan_id, "status": "hadir"})
    izin = await db.absensis.count_documents({"kegiatan_id": kegiatan_id, "status": "izin"})
    alpha = max(total - hadir - izin, 0)
    ratio = round((hadir / total) * 100, 1) if total else 0.0
    return {"total": total, "hadir": hadir, "izin": izin, "alpha": alpha, "ratio": ratio}


async def auto_close_kegiatan():
    """Tutup otomatis kegiatan (per-kegiatan) saat jam selesai WITA sudah lewat."""
    now = now_wita()
    async for k in db.kegiatans.find({"status": "open"}):
        end = kegiatan_end_dt(k.get("date"), k.get("end_time"))
        if end and now >= end:
            await db.kegiatans.update_one(
                {"_id": k["_id"]},
                {"$set": {"status": "closed", "closed_at": now.isoformat(), "auto_closed": True}})
            await revoke_delegations_for_kegiatan(k["_id"], "Kegiatan ditutup otomatis")


async def auto_close_loop():
    while True:
        try:
            await auto_close_kegiatan()
        except Exception as exc:  # pragma: no cover
            logger.warning("auto_close_loop error: %s", exc)
        await asyncio.sleep(60)


# ------------------------- Kegiatan CRUD -------------------------
@api_router.post("/admin/kegiatan")
async def create_kegiatan(body: KegiatanInput, admin: dict = Depends(require_staff)):
    if body.type not in KEGIATAN_TYPES:
        raise HTTPException(status_code=400, detail="Jenis kegiatan tidak valid")
    try:
        base_date = datetime.strptime(body.date, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="Format tanggal harus YYYY-MM-DD")
    if kegiatan_end_dt(body.date, body.start_time) is None or kegiatan_end_dt(body.date, body.end_time) is None:
        raise HTTPException(status_code=400, detail="Format waktu harus HH:MM")

    occurrences = 4 if body.recurring else 1
    group_id = str(uuid.uuid4()) if body.recurring else None
    now_iso = datetime.now(timezone.utc).isoformat()
    docs = []
    for i in range(occurrences):
        d = base_date + timedelta(weeks=i)
        docs.append({
            "_id": str(uuid.uuid4()),
            "name": body.name.strip(),
            "type": body.type,
            "date": d.strftime("%Y-%m-%d"),
            "start_time": body.start_time,
            "end_time": body.end_time,
            "teacher": (body.teacher or "").strip() or None,
            "material": (body.material or "").strip() or None,
            "location": (body.location or "").strip() or None,
            "recurring": body.recurring,
            "recurring_group_id": group_id,
            "status": "open",
            "closed_at": None,
            "auto_closed": False,
            "share_token": None,
            "share_expires_at": None,
            "created_at": now_iso,
            "created_by": str(admin["_id"]),
        })
    await db.kegiatans.insert_many(docs)
    await log_activity(admin, "buat_kegiatan",
                       f"Membuat kegiatan '{body.name}'" + (f" (berulang {occurrences}x)" if body.recurring else ""))
    total = await count_peserta_total()
    return [serialize_kegiatan(d, await kegiatan_counts(d["_id"], total)) for d in docs]


@api_router.get("/admin/kegiatan")
async def list_kegiatan(admin: dict = Depends(require_staff),
                        month: str = "", date_from: str = "", date_to: str = ""):
    query = {}
    if month.strip():
        query["date"] = {"$regex": f"^{re.escape(month.strip())}"}
    elif date_from.strip() or date_to.strip():
        rng = {}
        if date_from.strip():
            rng["$gte"] = date_from.strip()
        if date_to.strip():
            rng["$lte"] = date_to.strip()
        query["date"] = rng
    kegiatans = await db.kegiatans.find(query).sort([("date", -1), ("start_time", -1)]).to_list(1000)
    total = await count_peserta_total()
    return [serialize_kegiatan(k, await kegiatan_counts(k["_id"], total)) for k in kegiatans]


@api_router.get("/admin/kegiatan/{kegiatan_id}")
async def get_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    total = await count_peserta_total()
    return serialize_kegiatan(k, await kegiatan_counts(k["_id"], total))


@api_router.patch("/admin/kegiatan/{kegiatan_id}")
async def update_kegiatan(kegiatan_id: str, body: KegiatanUpdate, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    updates = {}
    for field in ["name", "type", "date", "start_time", "end_time", "teacher", "material", "location"]:
        val = getattr(body, field)
        if val is not None:
            if field == "type" and val not in KEGIATAN_TYPES:
                raise HTTPException(status_code=400, detail="Jenis kegiatan tidak valid")
            updates[field] = val.strip() if isinstance(val, str) else val
    if updates:
        await db.kegiatans.update_one({"_id": kegiatan_id}, {"$set": updates})
        await log_activity(admin, "ubah_kegiatan", f"Mengubah kegiatan '{k.get('name')}'")
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    return serialize_kegiatan(k, await kegiatan_counts(kegiatan_id))


@api_router.delete("/admin/kegiatan/{kegiatan_id}")
async def delete_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    await db.kegiatans.delete_one({"_id": kegiatan_id})
    await db.absensis.delete_many({"kegiatan_id": kegiatan_id})
    await log_activity(admin, "hapus_kegiatan", f"Menghapus kegiatan '{k.get('name')}'")
    return {"message": "Kegiatan dihapus"}


@api_router.post("/admin/kegiatan/{kegiatan_id}/close")
async def close_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    await db.kegiatans.update_one({"_id": kegiatan_id}, {"$set": {
        "status": "closed", "closed_at": now_wita().isoformat(), "auto_closed": False}})
    await revoke_delegations_for_kegiatan(kegiatan_id, "Kegiatan diselesaikan")
    await log_activity(admin, "selesaikan_kegiatan", f"Menyelesaikan kegiatan '{k.get('name')}'")
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    return serialize_kegiatan(k, await kegiatan_counts(kegiatan_id))


@api_router.post("/admin/kegiatan/{kegiatan_id}/reopen")
async def reopen_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    await db.kegiatans.update_one({"_id": kegiatan_id}, {"$set": {
        "status": "open", "closed_at": None, "auto_closed": False}})
    await log_activity(admin, "buka_kegiatan", f"Membuka kembali kegiatan '{k.get('name')}' untuk absen susulan")
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    return serialize_kegiatan(k, await kegiatan_counts(kegiatan_id))


# ------------------------- Absensi -------------------------
@api_router.post("/admin/kegiatan/{kegiatan_id}/absen")
async def mark_absen(kegiatan_id: str, body: AbsenInput, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    if body.status not in ABSEN_STATUS:
        raise HTTPException(status_code=400, detail="Status absensi tidak valid")
    u = await db.users.find_one({"_id": ObjectId(body.user_id)}) if ObjectId.is_valid(body.user_id) else None
    if not u:
        raise HTTPException(status_code=404, detail="Peserta tidak ditemukan")
    arrival = now_wita().isoformat() if body.status == "hadir" else None
    await db.absensis.update_one(
        {"kegiatan_id": kegiatan_id, "user_id": body.user_id},
        {"$set": {
            "kegiatan_id": kegiatan_id, "user_id": body.user_id, "status": body.status,
            "arrival_time": arrival, "marked_by": admin.get("name"),
            "marked_by_id": str(admin["_id"]), "updated_at": now_wita().isoformat()}},
        upsert=True)
    return {"user_id": body.user_id, "status": body.status, "arrival_time": arrival}


@api_router.get("/admin/kegiatan/{kegiatan_id}/rekap")
async def rekap_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    peserta = await db.users.find(PESERTA_QUERY).sort("name", 1).to_list(5000)
    absens = await db.absensis.find({"kegiatan_id": kegiatan_id}).to_list(10000)
    amap = {a["user_id"]: a for a in absens}
    rows, hadir, izin, alpha = [], 0, 0, 0
    gender = {"L": {"hadir": 0, "total": 0}, "P": {"hadir": 0, "total": 0}}
    for p in peserta:
        pid = str(p["_id"])
        a = amap.get(pid)
        status = a["status"] if a else "alpha"
        g = _derive_gender(p) or "L"
        if g not in gender:
            g = "L"
        gender[g]["total"] += 1
        if status == "hadir":
            hadir += 1
            gender[g]["hadir"] += 1
        elif status == "izin":
            izin += 1
        else:
            alpha += 1
        rows.append({
            "user_id": pid, "name": p.get("name"), "gender": _derive_gender(p),
            "status": status,
            "arrival_time": a.get("arrival_time") if a else None,
            "marked_by": a.get("marked_by") if a else None,
        })
    total = len(peserta)
    return {
        "kegiatan": serialize_kegiatan(k),
        "counts": {"total": total, "hadir": hadir, "izin": izin, "alpha": alpha,
                   "ratio": round((hadir / total) * 100, 1) if total else 0.0},
        "gender": gender,
        "rows": rows,
    }


# ------------------------- QR + Share (public rekap) -------------------------
async def ensure_share(kegiatan_id: str) -> dict:
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    now = datetime.now(timezone.utc)
    token = k.get("share_token")
    exp = k.get("share_expires_at")
    need_new = not token or not exp
    if exp:
        try:
            if datetime.fromisoformat(exp) <= now:
                need_new = True
        except Exception:
            need_new = True
    if need_new:
        token = secrets.token_urlsafe(10)
        exp = (now + timedelta(days=SHARE_EXPIRE_DAYS)).isoformat()
        await db.kegiatans.update_one({"_id": kegiatan_id},
                                      {"$set": {"share_token": token, "share_expires_at": exp}})
    link = f"{FRONTEND_URL}/rekap/{token}"
    return {"token": token, "link": link, "expires_at": exp}


@api_router.post("/admin/kegiatan/{kegiatan_id}/share")
async def share_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    info = await ensure_share(kegiatan_id)
    return info


@api_router.get("/admin/kegiatan/{kegiatan_id}/qr")
async def qr_kegiatan(kegiatan_id: str, admin: dict = Depends(require_staff)):
    info = await ensure_share(kegiatan_id)
    return {"link": info["link"], "expires_at": info["expires_at"],
            "image": make_qr_data_url(info["link"])}


@api_router.get("/rekap/{token}")
async def public_rekap(token: str):
    k = await db.kegiatans.find_one({"share_token": token})
    if not k:
        raise HTTPException(status_code=404, detail="Tautan rekap tidak ditemukan")
    exp = k.get("share_expires_at")
    if exp:
        try:
            if datetime.fromisoformat(exp) <= datetime.now(timezone.utc):
                raise HTTPException(status_code=410, detail="Tautan rekap sudah kadaluarsa")
        except HTTPException:
            raise
        except Exception:
            pass
    peserta = await db.users.find(PESERTA_QUERY).sort("name", 1).to_list(5000)
    absens = await db.absensis.find({"kegiatan_id": k["_id"]}).to_list(10000)
    amap = {a["user_id"]: a for a in absens}
    rows, hadir, izin, alpha = [], 0, 0, 0
    gender = {"L": 0, "P": 0}
    for p in peserta:
        a = amap.get(str(p["_id"]))
        status = a["status"] if a else "alpha"
        if status == "hadir":
            hadir += 1
            g = _derive_gender(p)
            if g in gender:
                gender[g] += 1
        elif status == "izin":
            izin += 1
        else:
            alpha += 1
        rows.append({"name": p.get("name"), "status": status,
                     "arrival_time": a.get("arrival_time") if a else None})
    total = len(peserta)
    return {
        "name": k.get("name"), "type": k.get("type"), "location": k.get("location"),
        "date": k.get("date"), "start_time": k.get("start_time"), "end_time": k.get("end_time"),
        "teacher": k.get("teacher"), "material": k.get("material"),
        "counts": {"total": total, "hadir": hadir, "izin": izin, "alpha": alpha,
                   "ratio": round((hadir / total) * 100, 1) if total else 0.0},
        "gender": gender,
        "rows": rows,
    }


# ------------------------- QR Absen Mandiri + Kesan & Pesan -------------------------
async def ensure_absen_token(kegiatan_id: str) -> dict:
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    token = k.get("absen_token")
    if not token:
        token = secrets.token_urlsafe(10)
        await db.kegiatans.update_one({"_id": kegiatan_id}, {"$set": {"absen_token": token}})
    link = f"{FRONTEND_URL}/absen/{token}"
    return {"token": token, "link": link, "kegiatan": k}


@api_router.post("/admin/kegiatan/{kegiatan_id}/absen-qr")
async def absen_qr(kegiatan_id: str, admin: dict = Depends(require_staff)):
    info = await ensure_absen_token(kegiatan_id)
    return {"token": info["token"], "link": info["link"],
            "image": make_qr_data_url(info["link"])}


@api_router.get("/absen/{token}")
async def public_absen_info(token: str):
    k = await db.kegiatans.find_one({"absen_token": token})
    if not k:
        raise HTTPException(status_code=404, detail="Tautan absen tidak ditemukan")
    peserta = await db.users.find(PESERTA_QUERY).sort("name", 1).to_list(5000)
    kmap = {km["_id"]: km.get("name") async for km in db.kelompoks.find()}
    absens = await db.absensis.find({"kegiatan_id": k["_id"]}).to_list(10000)
    amap = {a["user_id"]: a for a in absens}
    rows = []
    for p in peserta:
        pid = str(p["_id"])
        a = amap.get(pid)
        rows.append({
            "id": pid, "name": p.get("name"),
            "kelompok_name": kmap.get(p.get("kelompok_id")) if p.get("kelompok_id") else None,
            "status": a["status"] if a else "alpha",
            "arrival_time": a.get("arrival_time") if a else None,
        })
    return {
        "kegiatan": {
            "name": k.get("name"), "type": k.get("type"), "location": k.get("location"),
            "date": k.get("date"), "start_time": k.get("start_time"),
            "end_time": k.get("end_time"), "teacher": k.get("teacher"),
            "material": k.get("material"), "status": k.get("status", "open"),
        },
        "peserta": rows,
    }


@api_router.post("/absen/{token}/mark")
async def public_absen_mark(token: str, body: SelfAbsenInput):
    k = await db.kegiatans.find_one({"absen_token": token})
    if not k:
        raise HTTPException(status_code=404, detail="Tautan absen tidak ditemukan")
    if k.get("status", "open") != "open":
        raise HTTPException(status_code=403, detail="Kegiatan sudah ditutup. Absen mandiri dinonaktifkan.")
    u = await db.users.find_one({"_id": ObjectId(body.user_id)}) if ObjectId.is_valid(body.user_id) else None
    if not u or "peserta" not in (u.get("roles") or []):
        raise HTTPException(status_code=404, detail="Nama peserta tidak ditemukan")
    pid = str(u["_id"])
    existing = await db.absensis.find_one({"kegiatan_id": k["_id"], "user_id": pid})
    if existing and existing.get("status") == "hadir":
        return {"name": u.get("name"), "status": "hadir",
                "arrival_time": existing.get("arrival_time"), "already": True}
    arrival = now_wita().isoformat()
    await db.absensis.update_one(
        {"kegiatan_id": k["_id"], "user_id": pid},
        {"$set": {"kegiatan_id": k["_id"], "user_id": pid, "status": "hadir",
                  "arrival_time": arrival, "marked_by": "Mandiri (QR)",
                  "marked_by_id": None, "updated_at": arrival}},
        upsert=True)
    return {"name": u.get("name"), "status": "hadir", "arrival_time": arrival, "already": False}


@api_router.post("/absen/{token}/feedback")
async def public_absen_feedback(token: str, body: FeedbackInput):
    k = await db.kegiatans.find_one({"absen_token": token})
    if not k:
        raise HTTPException(status_code=404, detail="Tautan absen tidak ditemukan")
    msg = (body.message or "").strip()
    if not msg:
        raise HTTPException(status_code=400, detail="Pesan tidak boleh kosong.")
    doc = {"_id": str(uuid.uuid4()), "kegiatan_id": k["_id"],
           "name": (body.name or "").strip() or "Anonim", "message": msg[:1000],
           "created_at": now_wita().isoformat()}
    await db.feedbacks.insert_one(doc)
    return {"message": "Alhamdulillah, jazakumullahu khoiro."}


@api_router.get("/admin/kegiatan/{kegiatan_id}/feedback")
async def admin_kegiatan_feedback(kegiatan_id: str, admin: dict = Depends(require_staff)):
    items = await db.feedbacks.find({"kegiatan_id": kegiatan_id}).sort("created_at", -1).to_list(1000)
    return [{"id": f["_id"], "name": f.get("name"), "message": f.get("message"),
             "created_at": f.get("created_at")} for f in items]


# ===========================================================================
# FASE 3 (Pengurus) & FASE 4 (Peserta)
# ===========================================================================
MUSY_CATEGORIES = ["4S", "tim7"]
MAX_PINNED = 3
PERSONAL_QR_ROTATE = 60  # detik


def wa_number(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return None
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    elif digits.startswith("62"):
        pass
    elif digits.startswith("8"):
        digits = "62" + digits
    return digits


# ------------------------- Musyawarah -------------------------
class MusyawarahInput(BaseModel):
    category: str
    date: Optional[str] = None
    content: Optional[str] = ""


class MusyawarahUpdate(BaseModel):
    date: Optional[str] = None
    content: Optional[str] = None


def serialize_musyawarah(m: dict) -> dict:
    return {"id": m["_id"], "category": m.get("category"), "date": m.get("date"),
            "content": m.get("content", ""), "created_by": m.get("created_by_name"),
            "created_at": m.get("created_at"), "updated_at": m.get("updated_at")}


@api_router.get("/staff/musyawarah")
async def list_musyawarah(category: str = "", staff: dict = Depends(require_staff)):
    q = {}
    if category in MUSY_CATEGORIES:
        q["category"] = category
    items = await db.musyawarahs.find(q).sort("date", -1).to_list(2000)
    return [serialize_musyawarah(m) for m in items]


@api_router.post("/staff/musyawarah")
async def create_musyawarah(body: MusyawarahInput, staff: dict = Depends(require_staff)):
    if body.category not in MUSY_CATEGORIES:
        raise HTTPException(status_code=400, detail="Kategori tidak valid")
    now = now_wita().isoformat()
    doc = {"_id": str(uuid.uuid4()), "category": body.category,
           "date": body.date or now_wita().strftime("%Y-%m-%d"),
           "content": body.content or "", "created_by_id": str(staff["_id"]),
           "created_by_name": staff.get("name"), "created_at": now, "updated_at": now}
    await db.musyawarahs.insert_one(doc)
    return serialize_musyawarah(doc)


@api_router.patch("/staff/musyawarah/{musy_id}")
async def update_musyawarah(musy_id: str, body: MusyawarahUpdate, staff: dict = Depends(require_staff)):
    m = await db.musyawarahs.find_one({"_id": musy_id})
    if not m:
        raise HTTPException(status_code=404, detail="Catatan tidak ditemukan")
    updates = {"updated_at": now_wita().isoformat()}
    if body.date is not None:
        updates["date"] = body.date
    if body.content is not None:
        updates["content"] = body.content
    await db.musyawarahs.update_one({"_id": musy_id}, {"$set": updates})
    m = await db.musyawarahs.find_one({"_id": musy_id})
    return serialize_musyawarah(m)


@api_router.delete("/staff/musyawarah/{musy_id}")
async def delete_musyawarah(musy_id: str, staff: dict = Depends(require_staff)):
    res = await db.musyawarahs.delete_one({"_id": musy_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Catatan tidak ditemukan")
    return {"ok": True}


@api_router.get("/staff/musyawarah/{musy_id}/pdf")
async def musyawarah_pdf(musy_id: str, staff: dict = Depends(require_staff)):
    m = await db.musyawarahs.find_one({"_id": musy_id})
    if not m:
        raise HTTPException(status_code=404, detail="Catatan tidak ditemukan")
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
                            leftMargin=2 * cm, rightMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor="#0D5C3A")
    cat_label = "Musyawarah 4S" if m.get("category") == "4S" else "Musyawarah Tim 7"
    body_style = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, leading=17)
    elems = [Paragraph("E-KERTALANGU", title_style),
             Paragraph(cat_label, styles["Heading2"]),
             Paragraph(f"Tanggal: {m.get('date', '-')}", styles["Normal"]),
             Spacer(1, 0.4 * cm)]
    content = (m.get("content") or "").replace("\n", "<br/>")
    elems.append(Paragraph(content or "(kosong)", body_style))
    doc.build(elems)
    buf.seek(0)
    fname = f"musyawarah_{m.get('category')}_{m.get('date')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@api_router.get("/staff/musyawarah-export-pdf")
async def musyawarah_export_pdf(category: str = "", date_from: str = "", date_to: str = "",
                                staff: dict = Depends(require_staff)):
    q = {}
    if category in MUSY_CATEGORIES:
        q["category"] = category
    if date_from or date_to:
        dr = {}
        if date_from:
            dr["$gte"] = date_from
        if date_to:
            dr["$lte"] = date_to
        q["date"] = dr
    items = await db.musyawarahs.find(q).sort([("category", 1), ("date", 1)]).to_list(5000)
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
                            leftMargin=2 * cm, rightMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor="#0D5C3A")
    body_style = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, leading=17)
    cat_label = ("Musyawarah 4S" if category == "4S" else
                 "Musyawarah Tim 7" if category == "tim7" else "Musyawarah 4S & Tim 7")
    periode = ""
    if date_from or date_to:
        periode = f"Periode: {date_from or '...'} s/d {date_to or '...'}"
    elems = [Paragraph("E-KERTALANGU", title_style),
             Paragraph(f"Rekap {cat_label}", styles["Heading2"])]
    if periode:
        elems.append(Paragraph(periode, styles["Normal"]))
    elems.append(Spacer(1, 0.4 * cm))
    if not items:
        elems.append(Paragraph("Tidak ada catatan pada periode ini.", body_style))
    for m in items:
        lbl = "4S" if m.get("category") == "4S" else "Tim 7"
        elems.append(Paragraph(f"<b>[{lbl}] {m.get('date', '-')}</b>", styles["Heading3"]))
        content = (m.get("content") or "(kosong)").replace("\n", "<br/>")
        elems.append(Paragraph(content, body_style))
        elems.append(Spacer(1, 0.25 * cm))
        elems.append(HRFlowable(width="100%", thickness=0.5, color="#D1D5DB"))
        elems.append(Spacer(1, 0.25 * cm))
    doc.build(elems)
    buf.seek(0)
    tag = category if category in MUSY_CATEGORIES else "semua"
    fname = f"rekap_musyawarah_{tag}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
class PengumumanInput(BaseModel):
    title: str
    body: str = ""
    kegiatan_id: Optional[str] = None
    pengajar: Optional[str] = None
    important: bool = False
    pinned: bool = False
    pin_roles: List[str] = []


class PengumumanUpdate(BaseModel):
    title: Optional[str] = None
    body: Optional[str] = None
    kegiatan_id: Optional[str] = None
    pengajar: Optional[str] = None
    important: Optional[bool] = None
    pinned: Optional[bool] = None
    pin_roles: Optional[List[str]] = None


def serialize_pengumuman(p: dict) -> dict:
    return {"id": p["_id"], "title": p.get("title"), "body": p.get("body", ""),
            "kegiatan_id": p.get("kegiatan_id"), "kegiatan_name": p.get("kegiatan_name"),
            "pengajar": p.get("pengajar"), "important": bool(p.get("important")),
            "pinned": bool(p.get("pinned")), "pin_roles": p.get("pin_roles", []),
            "created_by": p.get("created_by_name"), "created_at": p.get("created_at")}


async def _count_pinned(exclude_id: Optional[str] = None) -> int:
    q = {"pinned": True}
    if exclude_id:
        q["_id"] = {"$ne": exclude_id}
    return await db.pengumumans.count_documents(q)


@api_router.get("/staff/pengumuman")
async def list_pengumuman(staff: dict = Depends(require_staff)):
    items = await db.pengumumans.find().sort([("pinned", -1), ("created_at", -1)]).to_list(2000)
    return [serialize_pengumuman(p) for p in items]


@api_router.post("/staff/pengumuman")
async def create_pengumuman(body: PengumumanInput, staff: dict = Depends(require_staff)):
    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Judul wajib diisi")
    if body.pinned and await _count_pinned() >= MAX_PINNED:
        raise HTTPException(status_code=400, detail=f"Maksimal {MAX_PINNED} pengumuman yang bisa di-pin")
    pin_roles = [r for r in (body.pin_roles or []) if r in VALID_ROLES]
    keg_name = None
    if body.kegiatan_id:
        k = await db.kegiatans.find_one({"_id": body.kegiatan_id})
        keg_name = k.get("name") if k else None
    doc = {"_id": str(uuid.uuid4()), "title": title, "body": body.body or "",
           "kegiatan_id": body.kegiatan_id, "kegiatan_name": keg_name,
           "pengajar": body.pengajar, "important": bool(body.important),
           "pinned": bool(body.pinned), "pin_roles": pin_roles if body.pinned else [],
           "created_by_id": str(staff["_id"]), "created_by_name": staff.get("name"),
           "created_at": now_wita().isoformat()}
    await db.pengumumans.insert_one(doc)
    await log_activity(staff, "buat_pengumuman", f"Membuat pengumuman '{title}'")
    return serialize_pengumuman(doc)


@api_router.patch("/staff/pengumuman/{peng_id}")
async def update_pengumuman(peng_id: str, body: PengumumanUpdate, staff: dict = Depends(require_staff)):
    p = await db.pengumumans.find_one({"_id": peng_id})
    if not p:
        raise HTTPException(status_code=404, detail="Pengumuman tidak ditemukan")
    updates = {}
    want_pinned = body.pinned if body.pinned is not None else p.get("pinned")
    if body.pinned is True and not p.get("pinned"):
        if await _count_pinned(exclude_id=peng_id) >= MAX_PINNED:
            raise HTTPException(status_code=400, detail=f"Maksimal {MAX_PINNED} pengumuman yang bisa di-pin")
    for field in ["title", "body", "pengajar", "important"]:
        val = getattr(body, field)
        if val is not None:
            updates[field] = val
    if body.kegiatan_id is not None:
        updates["kegiatan_id"] = body.kegiatan_id
        k = await db.kegiatans.find_one({"_id": body.kegiatan_id}) if body.kegiatan_id else None
        updates["kegiatan_name"] = k.get("name") if k else None
    if body.pinned is not None:
        updates["pinned"] = bool(body.pinned)
    if body.pin_roles is not None:
        updates["pin_roles"] = [r for r in body.pin_roles if r in VALID_ROLES]
    if not want_pinned:
        updates["pin_roles"] = []
    await db.pengumumans.update_one({"_id": peng_id}, {"$set": updates})
    p = await db.pengumumans.find_one({"_id": peng_id})
    return serialize_pengumuman(p)


@api_router.delete("/staff/pengumuman/{peng_id}")
async def delete_pengumuman(peng_id: str, staff: dict = Depends(require_staff)):
    res = await db.pengumumans.delete_one({"_id": peng_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pengumuman tidak ditemukan")
    return {"ok": True}


@api_router.get("/me/announcements")
async def my_announcements(role: str = "", user: dict = Depends(get_current_user)):
    roles = user.get("roles", [])
    target = role if role in roles else (roles[0] if roles else "peserta")
    items = await db.pengumumans.find({"pinned": True, "pin_roles": target}) \
        .sort("created_at", -1).to_list(MAX_PINNED)
    return [serialize_pengumuman(p) for p in items[:MAX_PINNED]]


# ------------------------- Pengingat Kegiatan (WA) -------------------------
def build_reminder_text(k: dict) -> str:
    lines = [f"*Undangan Pengajian - {k.get('name', '')}*", ""]
    lines.append(f"Hari/Tanggal : {k.get('date', '-')}")
    lines.append(f"Waktu : {k.get('start_time', '')} - {k.get('end_time', '')} WITA")
    if k.get("location"):
        lines.append(f"Tempat : {k.get('location')}")
    if k.get("teacher"):
        lines.append(f"Pengajar : {k.get('teacher')}")
    if k.get("material"):
        lines.append(f"Materi : {k.get('material')}")
    lines.append("")
    lines.append("Mohon kehadiran dan doa restunya. Jazakumullahu khoiro.")
    return "\n".join(lines)


@api_router.get("/staff/kegiatan/{kegiatan_id}/reminder")
async def kegiatan_reminder(kegiatan_id: str, staff: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    peserta = await db.users.find(PESERTA_QUERY).sort("name", 1).to_list(5000)
    recipients = []
    for p in peserta:
        wa = wa_number(p.get("whatsapp") or p.get("phone"))
        if wa:
            recipients.append({"id": str(p["_id"]), "name": p.get("name"),
                               "phone": p.get("whatsapp") or p.get("phone"), "wa": wa})
    return {"text": build_reminder_text(k), "recipients": recipients}


# ------------------------- Delegasi Absensi -------------------------
class DelegationInput(BaseModel):
    grantee_id: str
    reason: Optional[str] = ""


def serialize_delegation(d: dict) -> dict:
    return {"id": d["_id"], "kegiatan_id": d.get("kegiatan_id"),
            "kegiatan_name": d.get("kegiatan_name"), "granted_by": d.get("granted_by_name"),
            "grantee_id": d.get("grantee_id"), "grantee_name": d.get("grantee_name"),
            "reason": d.get("reason"), "active": bool(d.get("active")),
            "created_at": d.get("created_at"), "revoked_at": d.get("revoked_at"),
            "revoked_reason": d.get("revoked_reason")}


async def revoke_delegations_for_kegiatan(kegiatan_id: str, reason: str = "Kegiatan ditutup"):
    now = now_wita().isoformat()
    await db.delegations.update_many(
        {"kegiatan_id": kegiatan_id, "active": True},
        {"$set": {"active": False, "revoked_at": now, "revoked_reason": reason}})


@api_router.post("/staff/kegiatan/{kegiatan_id}/delegate")
async def create_delegation(kegiatan_id: str, body: DelegationInput, staff: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    if k.get("status", "open") != "open":
        raise HTTPException(status_code=400, detail="Kegiatan sudah ditutup, tidak bisa delegasi")
    reason = (body.reason or "").strip()
    g = await db.users.find_one({"_id": ObjectId(body.grantee_id)}) if ObjectId.is_valid(body.grantee_id) else None
    if not g:
        raise HTTPException(status_code=404, detail="Penerima delegasi tidak ditemukan")
    # cabut delegasi aktif sebelumnya utk kegiatan+penerima yang sama
    await db.delegations.update_many(
        {"kegiatan_id": kegiatan_id, "grantee_id": body.grantee_id, "active": True},
        {"$set": {"active": False, "revoked_at": now_wita().isoformat(),
                  "revoked_reason": "Diganti delegasi baru"}})
    now = now_wita().isoformat()
    doc = {"_id": str(uuid.uuid4()), "kegiatan_id": kegiatan_id,
           "kegiatan_name": k.get("name"), "granted_by_id": str(staff["_id"]),
           "granted_by_name": staff.get("name"), "grantee_id": body.grantee_id,
           "grantee_name": g.get("name"), "reason": reason, "active": True,
           "created_at": now, "revoked_at": None, "revoked_reason": None}
    await db.delegations.insert_one(doc)
    await log_activity(staff, "delegasi_absen",
                       f"Delegasi absen '{k.get('name')}' kepada {g.get('name')} — alasan: {reason or '(tanpa catatan)'}",
                       body.grantee_id)
    return serialize_delegation(doc)


@api_router.get("/staff/kegiatan/{kegiatan_id}/delegations")
async def list_delegations(kegiatan_id: str, staff: dict = Depends(require_staff)):
    items = await db.delegations.find({"kegiatan_id": kegiatan_id}).sort("created_at", -1).to_list(500)
    return [serialize_delegation(d) for d in items]


@api_router.post("/staff/delegation/{deleg_id}/revoke")
async def revoke_delegation(deleg_id: str, staff: dict = Depends(require_staff)):
    d = await db.delegations.find_one({"_id": deleg_id})
    if not d:
        raise HTTPException(status_code=404, detail="Delegasi tidak ditemukan")
    await db.delegations.update_one({"_id": deleg_id}, {"$set": {
        "active": False, "revoked_at": now_wita().isoformat(),
        "revoked_reason": f"Dicabut oleh {staff.get('name')}"}})
    await log_activity(staff, "cabut_delegasi",
                       f"Mencabut delegasi absen '{d.get('kegiatan_name')}' dari {d.get('grantee_name')}",
                       d.get("grantee_id"))
    d = await db.delegations.find_one({"_id": deleg_id})
    return serialize_delegation(d)


async def _active_delegation(kegiatan_id: str, user_id: str):
    return await db.delegations.find_one(
        {"kegiatan_id": kegiatan_id, "grantee_id": user_id, "active": True})


@api_router.get("/me/delegations")
async def my_delegations(user: dict = Depends(get_current_user)):
    uid = str(user["_id"])
    items = await db.delegations.find({"grantee_id": uid, "active": True}) \
        .sort("created_at", -1).to_list(200)
    result = []
    for d in items:
        k = await db.kegiatans.find_one({"_id": d.get("kegiatan_id")})
        if not k or k.get("status", "open") != "open":
            continue
        out = serialize_delegation(d)
        out["kegiatan"] = serialize_kegiatan(k)
        result.append(out)
    return result


@api_router.get("/delegate/kegiatan/{kegiatan_id}")
async def delegate_kegiatan_info(kegiatan_id: str, user: dict = Depends(get_current_user)):
    deleg = await _active_delegation(kegiatan_id, str(user["_id"]))
    if not deleg:
        raise HTTPException(status_code=403, detail="Anda tidak memiliki hak delegasi untuk kegiatan ini")
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    peserta = await db.users.find(PESERTA_QUERY).sort("name", 1).to_list(5000)
    kmap = {km["_id"]: km.get("name") async for km in db.kelompoks.find()}
    absens = await db.absensis.find({"kegiatan_id": kegiatan_id}).to_list(10000)
    amap = {a["user_id"]: a for a in absens}
    rows = []
    for p in peserta:
        pid = str(p["_id"])
        a = amap.get(pid)
        rows.append({"id": pid, "name": p.get("name"),
                     "kelompok_name": kmap.get(p.get("kelompok_id")) if p.get("kelompok_id") else None,
                     "status": a["status"] if a else "alpha",
                     "arrival_time": a.get("arrival_time") if a else None})
    return {"kegiatan": serialize_kegiatan(k), "peserta": rows,
            "delegation": serialize_delegation(deleg)}


@api_router.post("/delegate/kegiatan/{kegiatan_id}/absen")
async def delegate_mark_absen(kegiatan_id: str, body: AbsenInput, user: dict = Depends(get_current_user)):
    deleg = await _active_delegation(kegiatan_id, str(user["_id"]))
    if not deleg:
        raise HTTPException(status_code=403, detail="Anda tidak memiliki hak delegasi untuk kegiatan ini")
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    if k.get("status", "open") != "open":
        raise HTTPException(status_code=403, detail="Kegiatan sudah ditutup")
    if body.status not in ABSEN_STATUS:
        raise HTTPException(status_code=400, detail="Status absensi tidak valid")
    u = await db.users.find_one({"_id": ObjectId(body.user_id)}) if ObjectId.is_valid(body.user_id) else None
    if not u:
        raise HTTPException(status_code=404, detail="Peserta tidak ditemukan")
    arrival = now_wita().isoformat() if body.status == "hadir" else None
    await db.absensis.update_one(
        {"kegiatan_id": kegiatan_id, "user_id": body.user_id},
        {"$set": {"kegiatan_id": kegiatan_id, "user_id": body.user_id, "status": body.status,
                  "arrival_time": arrival, "marked_by": f"Delegasi: {user.get('name')}",
                  "marked_by_id": str(user["_id"]), "updated_at": now_wita().isoformat()}},
        upsert=True)
    return {"user_id": body.user_id, "status": body.status, "arrival_time": arrival}


# ------------------------- Peserta: QR pribadi rotating -------------------------
def _current_window() -> int:
    return int(datetime.now(timezone.utc).timestamp() // PERSONAL_QR_ROTATE)


def make_personal_token(user_id: str, window: int) -> str:
    msg = f"{user_id}.{window}"
    sig = hmac.new(get_jwt_secret().encode(), msg.encode(), hashlib.sha256).hexdigest()[:16]
    raw = f"{user_id}.{window}.{sig}"
    return base64.urlsafe_b64encode(raw.encode()).decode()


def verify_personal_token(token: str) -> Optional[str]:
    try:
        raw = base64.urlsafe_b64decode(token.encode()).decode()
        user_id, window_s, sig = raw.split(".")
        window = int(window_s)
    except Exception:
        return None
    now_w = _current_window()
    if window not in (now_w, now_w - 1):
        return None
    expected = hmac.new(get_jwt_secret().encode(), f"{user_id}.{window}".encode(),
                        hashlib.sha256).hexdigest()[:16]
    if not hmac.compare_digest(expected, sig):
        return None
    return user_id


@api_router.get("/me/qr")
async def my_personal_qr(user: dict = Depends(get_current_user)):
    window = _current_window()
    token = make_personal_token(str(user["_id"]), window)
    content = f"EKP:{token}"
    now_ts = datetime.now(timezone.utc).timestamp()
    expires_in = int((window + 1) * PERSONAL_QR_ROTATE - now_ts)
    return {"content": content, "image": make_qr_data_url(content),
            "rotate_seconds": PERSONAL_QR_ROTATE, "expires_in": max(expires_in, 1)}


class ScanPersonalInput(BaseModel):
    content: str


@api_router.post("/staff/kegiatan/{kegiatan_id}/scan-personal")
async def scan_personal_qr(kegiatan_id: str, body: ScanPersonalInput, staff: dict = Depends(require_staff)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    if k.get("status", "open") != "open":
        raise HTTPException(status_code=403, detail="Kegiatan sudah ditutup")
    content = (body.content or "").strip()
    if content.startswith("EKP:"):
        content = content[4:]
    user_id = verify_personal_token(content)
    if not user_id:
        raise HTTPException(status_code=400, detail="QR pribadi tidak valid atau sudah kadaluarsa")
    u = await db.users.find_one({"_id": ObjectId(user_id)}) if ObjectId.is_valid(user_id) else None
    if not u or "peserta" not in (u.get("roles") or []):
        raise HTTPException(status_code=404, detail="Peserta tidak ditemukan")
    existing = await db.absensis.find_one({"kegiatan_id": kegiatan_id, "user_id": user_id})
    if existing and existing.get("status") == "hadir":
        return {"name": u.get("name"), "status": "hadir",
                "arrival_time": existing.get("arrival_time"), "already": True}
    arrival = now_wita().isoformat()
    await db.absensis.update_one(
        {"kegiatan_id": kegiatan_id, "user_id": user_id},
        {"$set": {"kegiatan_id": kegiatan_id, "user_id": user_id, "status": "hadir",
                  "arrival_time": arrival, "marked_by": f"Dibantu: {staff.get('name')}",
                  "marked_by_id": str(staff["_id"]), "updated_at": arrival}},
        upsert=True)
    return {"name": u.get("name"), "status": "hadir", "arrival_time": arrival, "already": False}


# ------------------------- Peserta: dashboard / kegiatan / profil -------------------------
async def _my_attendance_ratio(user_id: str) -> dict:
    today = now_wita().strftime("%Y-%m-%d")
    kids = await db.kegiatans.find({"date": {"$lte": today}}).to_list(10000)
    kid_set = [k["_id"] for k in kids]
    total = len(kid_set)
    hadir = 0
    if kid_set:
        hadir = await db.absensis.count_documents(
            {"user_id": user_id, "status": "hadir", "kegiatan_id": {"$in": kid_set}})
    ratio = round((hadir / total) * 100, 1) if total else 0.0
    return {"total": total, "hadir": hadir, "ratio": ratio}


@api_router.get("/me/attendance-history")
async def my_attendance_history(months: int = 6, user: dict = Depends(get_current_user)):
    uid = str(user["_id"])
    months = max(1, min(months, 12))
    now = now_wita()
    # build list of last N months (oldest -> newest)
    ym_list = []
    y, m = now.year, now.month
    for _ in range(months):
        ym_list.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    ym_list.reverse()
    id_label = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    today = now.strftime("%Y-%m-%d")
    result = []
    cur = {"hadir": 0, "izin": 0, "alpha": 0, "total": 0}
    for ym in ym_list:
        kids = await db.kegiatans.find({"date": {"$regex": f"^{ym}", "$lte": today}}).to_list(5000)
        kid_set = [k["_id"] for k in kids]
        total = len(kid_set)
        hadir = izin = 0
        if kid_set:
            hadir = await db.absensis.count_documents({"user_id": uid, "status": "hadir", "kegiatan_id": {"$in": kid_set}})
            izin = await db.absensis.count_documents({"user_id": uid, "status": "izin", "kegiatan_id": {"$in": kid_set}})
        alpha = max(total - hadir - izin, 0)
        mi = int(ym[5:7]) - 1
        row = {"month": ym, "label": id_label[mi], "hadir": hadir, "izin": izin, "alpha": alpha, "total": total}
        result.append(row)
        if ym == ym_list[-1]:
            cur = {"hadir": hadir, "izin": izin, "alpha": alpha, "total": total}
    return {"months": result, "current": cur}


@api_router.get("/me/dashboard")
async def peserta_dashboard(user: dict = Depends(get_current_user)):
    uid = str(user["_id"])
    today = now_wita().strftime("%Y-%m-%d")
    upcoming_docs = await db.kegiatans.find({"date": {"$gte": today}}) \
        .sort([("date", 1), ("start_time", 1)]).to_list(20)
    upcoming = [serialize_kegiatan(k) for k in upcoming_docs[:5]]
    ann = await db.pengumumans.find({"pinned": True, "pin_roles": "peserta"}) \
        .sort("created_at", -1).to_list(MAX_PINNED)
    return {"name": user.get("name"), "gender": _derive_gender(user),
            "attendance": await _my_attendance_ratio(uid),
            "upcoming": upcoming,
            "announcements": [serialize_pengumuman(p) for p in ann[:MAX_PINNED]]}


@api_router.get("/me/kegiatan")
async def peserta_kegiatan(month: str = "", user: dict = Depends(get_current_user)):
    q = {}
    if month:
        q["date"] = {"$regex": f"^{month}"}
    docs = await db.kegiatans.find(q).sort([("date", -1), ("start_time", -1)]).to_list(2000)
    uid = str(user["_id"])
    absens = await db.absensis.find({"user_id": uid}).to_list(10000)
    amap = {a["kegiatan_id"]: a for a in absens}
    result = []
    for k in docs:
        item = serialize_kegiatan(k)
        a = amap.get(k["_id"])
        item["my_status"] = a["status"] if a else "alpha"
        item["my_arrival"] = a.get("arrival_time") if a else None
        result.append(item)
    return result


@api_router.get("/me/kegiatan/{kegiatan_id}")
async def peserta_kegiatan_detail(kegiatan_id: str, user: dict = Depends(get_current_user)):
    k = await db.kegiatans.find_one({"_id": kegiatan_id})
    if not k:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    uid = str(user["_id"])
    a = await db.absensis.find_one({"kegiatan_id": kegiatan_id, "user_id": uid})
    item = serialize_kegiatan(k)
    item["my_status"] = a["status"] if a else "alpha"
    item["my_arrival"] = a.get("arrival_time") if a else None
    return item


class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    dob: Optional[str] = None
    birthplace: Optional[str] = None
    address: Optional[str] = None
    gender: Optional[str] = None
    education: Optional[str] = None


@api_router.patch("/me/profile")
async def update_my_profile(body: ProfileUpdate, user: dict = Depends(get_current_user)):
    updates = {}
    for field in ["name", "phone", "whatsapp", "birthplace", "address", "education"]:
        val = getattr(body, field)
        if val is not None:
            updates[field] = val
    if body.dob is not None:
        updates["dob"] = normalize_dob(body.dob)
    if body.gender is not None:
        updates["gender"] = normalize_gender(body.gender)
    if updates:
        await db.users.update_one({"_id": user["_id"]}, {"$set": updates})
    u = await db.users.find_one({"_id": user["_id"]})
    return public_user(u)


# ------------------------- Dashboard -------------------------
@api_router.get("/admin/dashboard")
async def admin_dashboard(admin: dict = Depends(require_staff)):
    peserta = await db.users.find(PESERTA_QUERY).to_list(10000)
    total_peserta = len(peserta)
    lk = sum(1 for p in peserta if _derive_gender(p) == "L")
    pr = sum(1 for p in peserta if _derive_gender(p) == "P")
    akun_aktif = await db.users.count_documents({"status": "active"})
    akun_nonaktif = await db.users.count_documents({"status": {"$ne": "active"}})

    now = now_wita()
    this_month = now.strftime("%Y-%m")
    keg_this_month = await db.kegiatans.find({"date": {"$regex": f"^{this_month}"}}).to_list(1000)
    kegiatan_bulan_ini = len(keg_this_month)

    # rasio kehadiran bulan ini (rata-rata per kegiatan)
    ratios = []
    for k in keg_this_month:
        c = await kegiatan_counts(k["_id"], total_peserta)
        if c["total"]:
            ratios.append(c["ratio"])
    rasio_bulan = round(sum(ratios) / len(ratios), 1) if ratios else 0.0

    # tren 6 bulan terakhir
    tren = []
    for i in range(5, -1, -1):
        m = (now.replace(day=1) - timedelta(days=1)).replace(day=1) if False else None
        # hitung bulan target
        year = now.year
        month = now.month - i
        while month <= 0:
            month += 12
            year -= 1
        key = f"{year:04d}-{month:02d}"
        kegs = await db.kegiatans.find({"date": {"$regex": f"^{key}"}}).to_list(1000)
        rs = []
        for k in kegs:
            c = await kegiatan_counts(k["_id"], total_peserta)
            if c["total"]:
                rs.append(c["ratio"])
        tren.append({"month": key, "ratio": round(sum(rs) / len(rs), 1) if rs else 0.0,
                     "kegiatan": len(kegs)})

    today = now.strftime("%Y-%m-%d")
    upcoming = await db.kegiatans.find({"date": {"$gte": today}}).sort([("date", 1), ("start_time", 1)]).to_list(5)
    recent = await db.kegiatans.find({"date": {"$lt": today}}).sort([("date", -1)]).to_list(5)
    tot = total_peserta
    return {
        "total_peserta": total_peserta, "peserta_L": lk, "peserta_P": pr,
        "akun_aktif": akun_aktif, "akun_nonaktif": akun_nonaktif,
        "kegiatan_bulan_ini": kegiatan_bulan_ini,
        "rasio_kehadiran_bulan": rasio_bulan,
        "donut": {"L": lk, "P": pr},
        "tren": tren,
        "upcoming": [serialize_kegiatan(k, await kegiatan_counts(k["_id"], tot)) for k in upcoming],
        "recent": [serialize_kegiatan(k, await kegiatan_counts(k["_id"], tot)) for k in recent],
    }


# ------------------------- Laporan + Export -------------------------
async def build_laporan(date_from: str, date_to: str) -> dict:
    if not date_from or not date_to:
        now = now_wita()
        first = now.replace(day=1).strftime("%Y-%m-%d")
        date_from = date_from or first
        date_to = date_to or now.strftime("%Y-%m-%d")
    query = {"date": {"$gte": date_from, "$lte": date_to}}
    kegiatans = await db.kegiatans.find(query).sort([("date", 1)]).to_list(2000)
    keg_ids = [k["_id"] for k in kegiatans]
    peserta = await db.users.find(PESERTA_QUERY).to_list(10000)
    total_peserta = len(peserta)
    absens = await db.absensis.find({"kegiatan_id": {"$in": keg_ids}}).to_list(100000) if keg_ids else []

    # per-user tally
    tally = {str(p["_id"]): {"name": p.get("name"), "gender": _derive_gender(p),
                             "hadir": 0, "izin": 0, "alpha": 0} for p in peserta}
    per_keg_status = {}  # kegiatan_id -> {uid: status}
    for a in absens:
        per_keg_status.setdefault(a["kegiatan_id"], {})[a["user_id"]] = a["status"]

    total_hadir = total_izin = total_alpha = 0
    gender_hadir = {"L": 0, "P": 0}
    rows = []
    for k in kegiatans:
        statuses = per_keg_status.get(k["_id"], {})
        h = sum(1 for s in statuses.values() if s == "hadir")
        iz = sum(1 for s in statuses.values() if s == "izin")
        al = max(total_peserta - h - iz, 0)
        total_hadir += h
        total_izin += iz
        total_alpha += al
        rows.append({"id": k["_id"], "name": k.get("name"), "date": k.get("date"),
                     "type": k.get("type"), "hadir": h, "izin": iz, "alpha": al,
                     "ratio": round((h / total_peserta) * 100, 1) if total_peserta else 0.0})
        # per-user tally
        for p in peserta:
            uid = str(p["_id"])
            s = statuses.get(uid, "alpha")
            tally[uid][s] += 1
            if s == "hadir":
                g = _derive_gender(p)
                if g in gender_hadir:
                    gender_hadir[g] += 1

    tvals = list(tally.values())
    top_rajin = sorted(tvals, key=lambda x: x["hadir"], reverse=True)[:5]
    top_alpha = sorted(tvals, key=lambda x: x["alpha"], reverse=True)[:5]
    n_keg = len(kegiatans)
    denom = n_keg * total_peserta
    return {
        "date_from": date_from, "date_to": date_to,
        "total_kegiatan": n_keg, "total_peserta": total_peserta,
        "summary": {
            "hadir": total_hadir, "izin": total_izin, "alpha": total_alpha,
            "ratio": round((total_hadir / denom) * 100, 1) if denom else 0.0,
        },
        "gender_hadir": gender_hadir,
        "per_kegiatan": rows,
        "top_rajin": top_rajin, "top_alpha": top_alpha,
    }


@api_router.get("/admin/laporan")
async def admin_laporan(admin: dict = Depends(require_staff), date_from: str = "", date_to: str = ""):
    return await build_laporan(date_from.strip(), date_to.strip())


@api_router.get("/admin/laporan/export")
async def admin_laporan_export(admin: dict = Depends(require_staff),
                               date_from: str = "", date_to: str = "", format: str = "excel"):
    data = await build_laporan(date_from.strip(), date_to.strip())
    fname_base = f"laporan_kehadiran_{data['date_from']}_sd_{data['date_to']}"
    if format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        elems = []
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
        if os.path.exists(logo_path):
            _lw = 3.6 * cm
            logo_img = RLImage(logo_path, width=_lw, height=_lw * 452 / 467)
            logo_img.hAlign = "CENTER"
            elems.append(logo_img)
            elems.append(Spacer(1, 8))
        elems += [Paragraph("Laporan Kehadiran — E-KERTALANGU", styles["Title"]),
                 Paragraph(f"Periode: {data['date_from']} s/d {data['date_to']}", styles["Normal"]),
                 Paragraph(f"Total Kegiatan: {data['total_kegiatan']} · Total Peserta: {data['total_peserta']} · "
                           f"Kehadiran: {data['summary']['ratio']}%", styles["Normal"]),
                 Spacer(1, 12)]
        tdata = [["Tanggal", "Kegiatan", "Hadir", "Izin", "Alpha", "%"]]
        for r in data["per_kegiatan"]:
            tdata.append([r["date"], r["name"], r["hadir"], r["izin"], r["alpha"], f"{r['ratio']}%"])
        t = Table(tdata, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D5C3A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F2")]),
        ]))
        elems.append(t)
        doc.build(elems)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f'attachment; filename="{fname_base}.pdf"'})

    # default: Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Per Kegiatan"
    ws.append(["Tanggal", "Kegiatan", "Jenis", "Hadir", "Izin", "Alpha", "Kehadiran %"])
    for r in data["per_kegiatan"]:
        ws.append([r["date"], r["name"], r["type"], r["hadir"], r["izin"], r["alpha"], r["ratio"]])
    ws2 = wb.create_sheet("Ringkasan")
    ws2.append(["Periode", f"{data['date_from']} s/d {data['date_to']}"])
    ws2.append(["Total Kegiatan", data["total_kegiatan"]])
    ws2.append(["Total Peserta", data["total_peserta"]])
    ws2.append(["Hadir", data["summary"]["hadir"]])
    ws2.append(["Izin", data["summary"]["izin"]])
    ws2.append(["Alpha", data["summary"]["alpha"]])
    ws2.append(["Kehadiran %", data["summary"]["ratio"]])
    ws2.append(["Hadir L", data["gender_hadir"]["L"]])
    ws2.append(["Hadir P", data["gender_hadir"]["P"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'})


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_URL, "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Startup: indexes + seed
# ---------------------------------------------------------------------------
async def seed_users():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "name": "Administrator", "email": admin_email, "username": "admin",
            "phone": "081100000001", "dob": "1990-01-01", "address": "Kantor Pusat",
            "password_hash": hash_password(admin_password),
            "roles": ["admin", "pengurus", "peserta"], "status": "active", "source": "seed",
            "avatar_gender": "male", "token_version": 0,
            "created_at": datetime.now(timezone.utc).isoformat()})
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_password)}})

    demos = [
        {"name": "Pak Pengurus", "email": "pengurus@ekertalangu.id", "username": "pengurus",
         "phone": "081200000002", "dob": "1985-05-10", "address": "Jl. Melati No. 2",
         "password": "Pengurus#2026", "roles": ["pengurus", "peserta"]},
        {"name": "Ibu Jamaah", "email": "peserta@ekertalangu.id", "username": "peserta",
         "phone": "081300000003", "dob": "1970-08-17", "address": "Jl. Mawar No. 3",
         "password": "Peserta#2026", "roles": ["peserta"]},
    ]
    for d in demos:
        if not await db.users.find_one({"email": d["email"]}):
            await db.users.insert_one({
                "name": d["name"], "email": d["email"], "username": d["username"],
                "phone": d["phone"], "dob": d["dob"], "address": d["address"],
                "password_hash": hash_password(d["password"]), "roles": d["roles"],
                "status": "active", "source": "seed", "avatar_gender": "female",
                "gender": "P",
                "token_version": 0, "created_at": datetime.now(timezone.utc).isoformat()})
    # Backfill gender for admin seed
    await db.users.update_one({"email": admin_email, "gender": {"$exists": False}},
                              {"$set": {"gender": "L"}})

async def seed_kelompok():
    defaults = ["Majelis Pusat", "Kelompok Timur", "Kelompok Barat"]
    for name in defaults:
        if not await db.kelompoks.find_one({"name": name}):
            await db.kelompoks.insert_one({
                "_id": str(uuid.uuid4()), "name": name, "description": None,
                "created_at": datetime.now(timezone.utc).isoformat()})

@app.on_event("startup")
async def startup():
    for idx in ["email_1", "username_1"]:
        try:
            await db.users.drop_index(idx)
        except Exception:
            pass
    await db.users.create_index("email", unique=True,
                                partialFilterExpression={"email": {"$type": "string"}})
    await db.users.create_index("username", unique=True,
                                partialFilterExpression={"username": {"$type": "string"}})
    await db.users.create_index("phone")
    await db.users.create_index("name")
    await db.users.create_index("status")
    await db.users.create_index("kelompok_id")
    await db.login_attempts.create_index("identifier")
    await db.login_attempts.create_index("email")
    await db.activity_logs.create_index("at")
    await db.kelompoks.create_index("name")
    await seed_users()
    await seed_kelompok()
    await get_or_create_public_qr()
    await db.kegiatans.create_index("date")
    await db.kegiatans.create_index("status")
    await db.kegiatans.create_index("share_token")
    await db.absensis.create_index([("kegiatan_id", 1), ("user_id", 1)], unique=True)
    await db.absensis.create_index("kegiatan_id")
    await db.absensis.create_index("user_id")
    await db.musyawarahs.create_index([("category", 1), ("date", -1)])
    await db.pengumumans.create_index("pinned")
    await db.pengumumans.create_index("pin_roles")
    await db.delegations.create_index([("kegiatan_id", 1), ("grantee_id", 1), ("active", 1)])
    await db.delegations.create_index("grantee_id")
    asyncio.create_task(auto_close_loop())

@app.on_event("shutdown")
async def shutdown():
    client.close()
